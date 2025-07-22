import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

class TestCaseExcelLogger:
    def __init__(self, excel_path="qa/test_results.xlsx"):
        self.excel_path = excel_path
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        if not os.path.exists(excel_path):
            self._create_workbook()
        self.wb = load_workbook(excel_path)
        self.ws = self.wb.active

    def _create_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "TestResults"
        ws.append([
            "Test Case ID", "Scenario", "Step", "Timestamp", "Action", "Verdict", "Result", "Before Screenshot", "After Screenshot", "Bug ID"
        ])
        wb.save(self.excel_path)

    def log_step(self, test_case_id, scenario, step, action, verdict, result, before_path, after_path, bug_id=None):
        timestamp = datetime.utcnow().isoformat()
        self.ws.append([
            test_case_id, scenario, step, timestamp, action, verdict, result, before_path, after_path, bug_id or ""
        ])
        self.wb.save(self.excel_path)

    def log_summary(self, test_case_id, scenario, total_steps, passed, failed, bug_count):
        # Add a summary row at the end of the test case
        self.ws.append([
            test_case_id, scenario, "SUMMARY", "", f"Total Steps: {total_steps}", f"Passed: {passed}", f"Failed: {failed}", "", "", f"Bugs: {bug_count}"
        ])
        self.wb.save(self.excel_path)

    def close(self):
        self.wb.save(self.excel_path)
        self.wb.close() 